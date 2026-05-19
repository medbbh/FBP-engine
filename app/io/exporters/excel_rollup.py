from io import BytesIO

import openpyxl
from openpyxl.styles import Font, PatternFill


def generate_national_rollup(national_data: dict) -> bytes:
    wb = openpyxl.Workbook()
    summary_ws = wb.active
    summary_ws.title = "Résumé national"

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(fill_type="solid", fgColor="2563EB")

    headers = ["Région", "Nb structures", "Total net (MRU)"]
    for col, h in enumerate(headers, start=1):
        cell = summary_ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill

    for row_num, (region, data) in enumerate(national_data.get("regions", {}).items(), start=2):
        summary_ws.cell(row=row_num, column=1, value=region)
        summary_ws.cell(row=row_num, column=2, value=data["facilities"])
        summary_ws.cell(row=row_num, column=3, value=float(data["total_net_amount"]))

    for region, data in national_data.get("regions", {}).items():
        ws = wb.create_sheet(title=region[:31])
        detail_headers = ["Structure", "Statut", "Montant net (MRU)"]
        for col, h in enumerate(detail_headers, start=1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.font = header_font
            cell.fill = header_fill
        for row_num, decl in enumerate(data["declarations"], start=2):
            ws.cell(row=row_num, column=1, value=decl["facility_name"])
            ws.cell(row=row_num, column=2, value=decl["status"])
            ws.cell(row=row_num, column=3, value=float(decl["net_amount"]) if decl["net_amount"] else None)

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()
